"""
report_export.py
~~~~~~~~~~~~~~~~
Generate downloadable Excel (.xlsx) and PDF reports from runway
calculation results.

Dependencies:
    - openpyxl  (Excel generation)
    - fpdf2     (PDF generation)
"""

from io import BytesIO
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

from fpdf import FPDF


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _fmt_currency(value):
    """Format a float as accounting-style currency.

    Positive:  "$1,234,567.89"
    Negative:  "($1,234,567.89)"
    Zero/None: "$0.00"
    """
    if value is None:
        return "$0.00"
    abs_val = abs(value)
    formatted = f"${abs_val:,.2f}"
    if value < 0:
        return f"({formatted})"
    return formatted


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def generate_excel(results):
    """Return a BytesIO object containing an .xlsx workbook built from
    *results* (the dict passed to the template by the runway calculator).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Runway Report"

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 5      # narrow; only needed for merge

    # Reusable styles
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=16)
    section_font = Font(bold=True, underline="single", size=12)
    section_font_no_ul = Font(bold=True, size=12)
    acct_fmt = '#,##0.00'

    def _write_label_value(row, label, value, bold=False, fmt=None):
        """Write a label in col-A and a value in col-B."""
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_b = ws.cell(row=row, column=2, value=value)
        if bold:
            cell_a.font = bold_font
            cell_b.font = bold_font
        if fmt:
            cell_b.number_format = fmt

    # ---- Row 1: Title (merged A1:C1) ------------------------------------
    ws.merge_cells("A1:C1")
    title_cell = ws.cell(row=1, column=1, value="Company Runway Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left")

    # ---- Row 2: blank ----------------------------------------------------

    # ---- Row 3: Report date ---------------------------------------------
    _write_label_value(3, "Report Date", date.today().strftime("%Y-%m-%d"))

    # ---- Row 4: blank ----------------------------------------------------

    # ---- Row 5: Balance Sheet Data header --------------------------------
    hdr5 = ws.cell(row=5, column=1, value="Balance Sheet Data")
    hdr5.font = section_font

    _write_label_value(6, "Balance Sheet Date", results.get("balance_sheet_date", ""))
    _write_label_value(7, "Cash & Cash Equivalents", results.get("cash_and_equivalents", 0), fmt=acct_fmt)
    _write_label_value(8, "Accounts Receivable", results.get("accounts_receivable", 0), fmt=acct_fmt)
    _write_label_value(9, "Total Liquid Assets", results.get("total_liquid_assets", 0), bold=True, fmt=acct_fmt)

    # ---- Row 10: blank ---------------------------------------------------

    # ---- Row 11: Income Statement header ---------------------------------
    quarter_used = results.get("quarter_used", "")
    hdr11 = ws.cell(row=11, column=1, value=f"Income Statement Data ({quarter_used})")
    hdr11.font = section_font_no_ul

    _write_label_value(12, "Quarterly Revenue", results.get("quarterly_revenue", 0), fmt=acct_fmt)
    _write_label_value(13, "Quarterly COGS", results.get("quarterly_cogs", 0), fmt=acct_fmt)
    _write_label_value(14, "Quarterly Operating Expenses", results.get("quarterly_opex", 0), fmt=acct_fmt)
    _write_label_value(15, "Quarterly Other Income", results.get("quarterly_other_income", 0), fmt=acct_fmt)
    _write_label_value(16, "Quarterly Other Expense", results.get("quarterly_other_expense", 0), fmt=acct_fmt)

    # ---- Row 17: blank ---------------------------------------------------

    # ---- Row 18: Monthly Averages header ---------------------------------
    hdr18 = ws.cell(row=18, column=1, value="Monthly Averages (Quarterly ÷ 3)")
    hdr18.font = section_font_no_ul

    _write_label_value(19, "Monthly Revenue", results.get("monthly_revenue", 0), fmt=acct_fmt)
    _write_label_value(20, "Monthly COGS", results.get("monthly_cogs", 0), fmt=acct_fmt)
    _write_label_value(21, "Monthly Operating Expenses", results.get("monthly_opex", 0), fmt=acct_fmt)
    _write_label_value(22, "Monthly Other Income", results.get("monthly_other_income", 0), fmt=acct_fmt)
    _write_label_value(23, "Monthly Other Expense", results.get("monthly_other_expense", 0), fmt=acct_fmt)

    # ---- Row 24: blank ---------------------------------------------------

    # ---- Row 25: Runway Analysis header ----------------------------------
    hdr25 = ws.cell(row=25, column=1, value="Runway Analysis")
    hdr25.font = section_font_no_ul

    _write_label_value(26, "Gross Monthly Burn", results.get("gross_monthly_burn", 0), fmt=acct_fmt)

    # Net burn – special handling for cash-flow-positive
    is_positive = results.get("is_cash_flow_positive", False)
    net_burn = results.get("net_monthly_burn", 0)
    if is_positive:
        _write_label_value(27, "Net Monthly Burn", "Cash Flow Positive")
    else:
        _write_label_value(27, "Net Monthly Burn", net_burn, fmt=acct_fmt)

    # Gross runway
    gross_runway = results.get("gross_runway_months")
    _write_label_value(28, "Gross Runway (months)",
                       round(gross_runway, 1) if gross_runway is not None else "N/A")

    # Net runway
    net_runway = results.get("net_runway_months")
    if is_positive:
        net_display = "Unlimited"
    elif net_runway is None:
        net_display = "N/A"
    else:
        net_display = round(net_runway, 1)
    _write_label_value(29, "Net Runway (months)", net_display)

    # Cash-out date
    end_date = results.get("runway_end_date")
    _write_label_value(30, "Projected Cash-Out Date", end_date if end_date else "N/A")

    # ---- Adjustments (if present) ----------------------------------------
    current_row = 30  # last written row

    adjustments = results.get("adjustments") or []
    adjusted = results.get("adjusted")

    if adjustments:
        current_row += 2  # skip a blank row
        ws.cell(row=current_row, column=1,
                value="One-Off Adjustments").font = section_font
        current_row += 1

        for adj in adjustments:
            adj_type = adj.get("type", "").replace("_", " ").title()
            cat_or_desc = adj.get("description") or adj.get("category", "")
            amount = adj.get("amount", 0)
            _write_label_value(current_row, f"{adj_type} – {cat_or_desc}",
                               amount, fmt=acct_fmt)
            current_row += 1

        current_row += 1  # blank row after adjustment list

    if adjusted:
        ws.cell(row=current_row, column=1,
                value="Adjusted Runway Analysis").font = section_font
        current_row += 1

        # Gross Monthly Burn
        _write_label_value(current_row, "Gross Monthly Burn",
                           adjusted.get("gross_monthly_burn", 0), fmt=acct_fmt)
        current_row += 1

        # Net Monthly Burn
        adj_positive = adjusted.get("is_cash_flow_positive", False)
        if adj_positive:
            _write_label_value(current_row, "Net Monthly Burn",
                               "Cash Flow Positive")
        else:
            _write_label_value(current_row, "Net Monthly Burn",
                               adjusted.get("net_monthly_burn", 0), fmt=acct_fmt)
        current_row += 1

        # Gross Runway (months)
        adj_gross_rw = adjusted.get("gross_runway_months")
        _write_label_value(current_row, "Gross Runway (months)",
                           round(adj_gross_rw, 1) if adj_gross_rw is not None else "N/A")
        current_row += 1

        # Net Runway (months)
        adj_net_rw = adjusted.get("net_runway_months")
        if adj_positive:
            adj_net_display = "Unlimited"
        elif adj_net_rw is None:
            adj_net_display = "N/A"
        else:
            adj_net_display = round(adj_net_rw, 1)
        _write_label_value(current_row, "Net Runway (months)", adj_net_display)
        current_row += 1

        # Projected Cash-Out Date
        adj_end = adjusted.get("runway_end_date")
        if isinstance(adj_end, (date, datetime)):
            adj_end_str = adj_end.strftime("%Y-%m-%d")
        elif adj_end:
            adj_end_str = str(adj_end)
        else:
            adj_end_str = "N/A"
        _write_label_value(current_row, "Projected Cash-Out Date", adj_end_str)

    # ---- Write to BytesIO ------------------------------------------------
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF report
# ---------------------------------------------------------------------------

class _RunwayPDF(FPDF):
    """Thin FPDF subclass used only internally for header/footer."""

    def header(self):
        # intentionally empty – we draw the header manually on page 1
        pass

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")


def generate_pdf(results):
    """Return a BytesIO object containing a PDF report built from *results*."""
    pdf = _RunwayPDF(orientation="P", unit="mm", format="Letter")
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    navy = (26, 26, 46)       # #1a1a2e
    black = (0, 0, 0)
    grey = (100, 100, 100)
    white = (255, 255, 255)
    row_alt = (245, 245, 250)  # light background for alternating rows
    orange = (230, 126, 34)    # adjustment section accent

    page_w = pdf.w - 30       # usable width (Letter=215.9 minus 2×15 margins)
    col_a = page_w * 0.55
    col_b = page_w * 0.45

    # ------------------------------------------------------------------
    # Header area
    # ------------------------------------------------------------------
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(*navy)
    pdf.cell(0, 10, "Company Runway Report", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(*grey)
    pdf.cell(0, 6, f"Generated on {date.today().strftime('%Y-%m-%d')}",
             new_x="LMARGIN", new_y="NEXT")

    # Horizontal rule
    pdf.ln(2)
    y_line = pdf.get_y()
    pdf.set_draw_color(*navy)
    pdf.set_line_width(0.6)
    pdf.line(15, y_line, 15 + page_w, y_line)
    pdf.ln(6)

    # ------------------------------------------------------------------
    # Helpers local to PDF generation
    # ------------------------------------------------------------------
    def _section_header(title, color=None):
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(*(color or navy))
        pdf.cell(0, 8, title, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)

    row_idx = [0]  # mutable counter for alternating row colour

    def _table_row(label, value, bold=False, text_color=None):
        # alternating background
        if row_idx[0] % 2 == 0:
            pdf.set_fill_color(*row_alt)
            fill = True
        else:
            pdf.set_fill_color(*white)
            fill = True

        style = "B" if bold else ""
        pdf.set_font("Helvetica", style, 10)
        color = text_color or black
        pdf.set_text_color(*color)

        pdf.cell(col_a, 7, f"  {label}", border=0, fill=fill)
        pdf.cell(col_b, 7, str(value), border=0, fill=fill,
                 align="R", new_x="LMARGIN", new_y="NEXT")
        row_idx[0] += 1

    def _reset_rows():
        row_idx[0] = 0

    # ------------------------------------------------------------------
    # Section 1 – Balance Sheet Data
    # ------------------------------------------------------------------
    _section_header("Balance Sheet Data")
    _reset_rows()

    bs_date = results.get("balance_sheet_date", "N/A")
    _table_row("Balance Sheet Date", bs_date)
    _table_row("Cash & Cash Equivalents",
               _fmt_currency(results.get("cash_and_equivalents", 0)))
    _table_row("Accounts Receivable",
               _fmt_currency(results.get("accounts_receivable", 0)))
    _table_row("Total Liquid Assets",
               _fmt_currency(results.get("total_liquid_assets", 0)), bold=True)

    # ------------------------------------------------------------------
    # Section 2 – Income Statement
    # ------------------------------------------------------------------
    quarter_used = results.get("quarter_used", "")
    _section_header(f"Income Statement ({quarter_used})")
    _reset_rows()

    inc_rows = [
        ("Revenue",            "quarterly_revenue",       "monthly_revenue"),
        ("COGS",               "quarterly_cogs",          "monthly_cogs"),
        ("Operating Expenses", "quarterly_opex",          "monthly_opex"),
        ("Other Income",       "quarterly_other_income",  "monthly_other_income"),
        ("Other Expense",      "quarterly_other_expense", "monthly_other_expense"),
    ]

    # Sub-header for the two-column layout
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(*grey)
    label_w = col_a * 0.55
    qtr_w = col_a * 0.45
    pdf.cell(label_w, 6, "")
    pdf.cell(qtr_w, 6, "Quarterly", align="R")
    pdf.cell(col_b, 6, "Monthly", align="R", new_x="LMARGIN", new_y="NEXT")

    for (label, q_key, m_key) in inc_rows:
        q_val = _fmt_currency(results.get(q_key, 0))
        m_val = _fmt_currency(results.get(m_key, 0))

        if row_idx[0] % 2 == 0:
            pdf.set_fill_color(*row_alt)
        else:
            pdf.set_fill_color(*white)

        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*black)
        pdf.cell(label_w, 7, f"  {label}", fill=True)
        pdf.cell(qtr_w, 7, q_val, align="R", fill=True)
        pdf.cell(col_b, 7, m_val, align="R", fill=True,
                 new_x="LMARGIN", new_y="NEXT")
        row_idx[0] += 1

    # ------------------------------------------------------------------
    # Section 3 – Runway Analysis
    # ------------------------------------------------------------------
    _section_header("Runway Analysis")
    _reset_rows()

    is_positive = results.get("is_cash_flow_positive", False)

    # Colour-coding helper for runway months
    def _runway_color(months):
        """Return (display_str, rgb_tuple) based on months remaining."""
        if months is None:
            return ("N/A", black)
        try:
            m = float(months)
        except (TypeError, ValueError):
            return (str(months), black)
        if m < 6:
            color = (204, 0, 0)        # red
        elif m <= 12:
            color = (204, 136, 0)      # amber
        else:
            color = (0, 153, 51)       # green
        return (f"{m:.1f}", color)

    _table_row("Gross Monthly Burn",
               _fmt_currency(results.get("gross_monthly_burn", 0)))

    if is_positive:
        _table_row("Net Monthly Burn", "Cash Flow Positive",
                   text_color=(0, 153, 51))
    else:
        _table_row("Net Monthly Burn",
                   _fmt_currency(results.get("net_monthly_burn", 0)))

    # Gross runway
    gross_months = results.get("gross_runway_months")
    g_display, g_color = _runway_color(gross_months)
    _table_row("Gross Runway (months)", g_display, text_color=g_color)

    # Net runway
    net_months = results.get("net_runway_months")
    if is_positive:
        _table_row("Net Runway (months)", "Unlimited",
                   text_color=(0, 153, 51))
    else:
        n_display, n_color = _runway_color(net_months)
        _table_row("Net Runway (months)", n_display, text_color=n_color)

    # Cash-out date
    end_date = results.get("runway_end_date")
    if is_positive or not end_date:
        _table_row("Projected Cash-Out Date", "N/A")
    else:
        # colour-code date using net runway months for consistency
        _, d_color = _runway_color(net_months)
        _table_row("Projected Cash-Out Date", str(end_date), text_color=d_color)

    # ------------------------------------------------------------------
    # Section 4 – One-Off Adjustments (if present)
    # ------------------------------------------------------------------
    adjustments = results.get("adjustments") or []
    adjusted = results.get("adjusted")

    if adjustments:
        _section_header("One-Off Adjustments", color=orange)
        _reset_rows()

        adj_type_w = page_w * 0.22
        adj_desc_w = page_w * 0.48
        adj_amt_w = page_w * 0.30

        for adj in adjustments:
            adj_type = adj.get("type", "").replace("_", " ").title()
            cat_or_desc = adj.get("description") or adj.get("category", "")
            amount = adj.get("amount", 0)

            if row_idx[0] % 2 == 0:
                pdf.set_fill_color(*row_alt)
            else:
                pdf.set_fill_color(*white)

            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(*black)
            pdf.cell(adj_type_w, 7, f"  {adj_type}", fill=True)
            pdf.cell(adj_desc_w, 7, cat_or_desc, fill=True)
            pdf.cell(adj_amt_w, 7, _fmt_currency(amount), align="R",
                     fill=True, new_x="LMARGIN", new_y="NEXT")
            row_idx[0] += 1

    # ------------------------------------------------------------------
    # Section 5 – Adjusted Runway Analysis (if present)
    # ------------------------------------------------------------------
    if adjusted:
        _section_header("Adjusted Runway Analysis")
        _reset_rows()

        adj_positive = adjusted.get("is_cash_flow_positive", False)

        _table_row("Gross Monthly Burn",
                   _fmt_currency(adjusted.get("gross_monthly_burn", 0)))

        if adj_positive:
            _table_row("Net Monthly Burn", "Cash Flow Positive",
                       text_color=(0, 153, 51))
        else:
            _table_row("Net Monthly Burn",
                       _fmt_currency(adjusted.get("net_monthly_burn", 0)))

        # Adjusted gross runway
        adj_gross = adjusted.get("gross_runway_months")
        ag_display, ag_color = _runway_color(adj_gross)
        _table_row("Gross Runway (months)", ag_display, text_color=ag_color)

        # Adjusted net runway
        adj_net = adjusted.get("net_runway_months")
        if adj_positive:
            _table_row("Net Runway (months)", "Unlimited",
                       text_color=(0, 153, 51))
        else:
            an_display, an_color = _runway_color(adj_net)
            _table_row("Net Runway (months)", an_display, text_color=an_color)

        # Adjusted cash-out date
        adj_end = adjusted.get("runway_end_date")
        if adj_positive or not adj_end:
            _table_row("Projected Cash-Out Date", "N/A")
        else:
            if isinstance(adj_end, (date, datetime)):
                adj_end_str = adj_end.strftime("%Y-%m-%d")
            else:
                adj_end_str = str(adj_end)
            _, ad_color = _runway_color(adj_net)
            _table_row("Projected Cash-Out Date", adj_end_str,
                       text_color=ad_color)

    # ------------------------------------------------------------------
    # Write to BytesIO
    # ------------------------------------------------------------------
    buf = BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    return buf
